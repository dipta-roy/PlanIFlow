from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, 
                             QTableWidgetItem, QHeaderView, QLabel, QComboBox, 
                             QPushButton, QScrollArea, QGroupBox, QFileDialog, QMessageBox,
                             QDialog, QDialogButtonBox, QTextBrowser)
from PyQt6.QtCore import Qt
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from data_manager.evm_calculator import EVMCalculator
import pandas as pd
import os

class EVMHelpDialog(QDialog):
    """Help dialog explaining Earned Value Management concepts."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("About Earned Value Management (EVM)")
        self.resize(700, 600)
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        text_browser = QTextBrowser()
        text_browser.setHtml("""
            <h1>Earned Value Management (EVM)</h1>
            
            <p>
            Earned Value Management (EVM) is a project management technique for measuring project performance and progress. 
            It integrates project scope, cost, and schedule to help project managers assess performance against the baseline.
            </p>
            
            <h2>Core Metrics</h2>
            <ul>
                <li><b>PV (Planned Value):</b> The budgeted cost for the work scheduled to be completed by the status date. (Also known as BCWS).</li>
                <li><b>EV (Earned Value):</b> The budgeted cost for the work <i>actually</i> performed. Calculated as: % Complete × BAC. (Also known as BCWP).</li>
                <li><b>AC (Actual Cost):</b> The actual cost incurred for the work performed so far. (Also known as ACWP).</li>
                <li><b>BAC (Budget at Completion):</b> The total planned cost of the project/task when completed.</li>
            </ul>
            
            <h2>Performance Indicators</h2>
            <ul>
                <li><b>CV (Cost Variance):</b> EV - AC. Positive is under budget, negative is over budget.</li>
                <li><b>SV (Schedule Variance):</b> EV - PV. Positive is ahead of schedule, negative is behind schedule.</li>
                <li><b>CPI (Cost Performance Index):</b> EV / AC. Value > 1.0 means project is performing efficiently regarding cost ($1 worth of work for every $1 spent).</li>
                <li><b>SPI (Schedule Performance Index):</b> EV / PV. Value > 1.0 means project is performing efficiently regarding time (ahead of schedule).</li>
            </ul>
            
            <h2>Forecasting</h2>
            <ul>
                <li><b>EAC (Estimate at Completion):</b> BAC / CPI. Predicted total cost of the project based on current performance.</li>
                <li><b>VAC (Variance at Completion):</b> BAC - EAC. Difference between original budget and the current estimate.</li>
            </ul>

            <hr/>
            <h2>Example Scenario</h2>
            <p>
            You have a task with a budget (BAC) of <b>$2,000</b> and a duration of 10 days. 
            On Day 5 (Status Date):
            </p>
            <ul>
                <li><b>Planned Value (PV):</b> You should be 50% done. PV = $1,000.</li>
                <li><b>Earned Value (EV):</b> You are actually only 40% done. EV = $2,000 × 0.40 = <b>$800</b>.</li>
                <li><b>Actual Cost (AC):</b> You have spent <b>$900</b> so far.</li>
            </ul>
            <p><b>Results:</b></p>
            <ul>
                <li><b>CV = $800 - $900 = -$100</b> (Over budget)</li>
                <li><b>SV = $800 - $1,000 = -$200</b> (Behind schedule)</li>
                <li><b>CPI = 0.89</b> (Getting $0.89 value for every $1 spent)</li>
                <li><b>SPI = 0.80</b> (Progressing at 80% of the planned rate)</li>
            </ul>

            <hr/>
            <h2>How to use in PlanIFlow</h2>
            <ol>
                <li><b>Create a Plan:</b> Define your tasks and assign resources with billing rates (this establishes Cost).</li>
                <li><b>Set a Baseline:</b> Go to <i>Settings &gt; Baselines &gt; Set Baseline</i>. This saves a snapshot of your planned costs and schedule (Calculation of PV).</li>
                <li><b>Execute & Track:</b> As real work happens:
                    <ul>
                        <li>Update the <b>% Complete</b> on tasks to reflect progress.</li>
                        <li>If tasks are delayed, move their dates—this affects AC (Actual Cost) vs the Baseline.</li>
                    </ul>
                </li>
                <li><b>View Analysis:</b> 
                    <ul>
                        <li>Open this <b>EVM Analysis</b> tab.</li>
                        <li>Select your saved <b>Baseline</b> from the dropdown.</li>
                        <li>All metrics (CV, SV, CPI, etc.) are calculated automatically comparing your <i>Current Data</i> against the <i>Selected Baseline</i>.</li>
                    </ul>
                </li>
            </ol>
        """)
        layout.addWidget(text_browser)
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        buttons.rejected.connect(self.accept)
        layout.addWidget(buttons)

class EVMAnalysisTab(QWidget):
    """
    UI Tab for Earned Value Management (EVM) Analysis.
    Displays metrics table, summary indicators, and performance charts.
    """
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.calculator = EVMCalculator(main_window.data_manager)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # --- Header Section ---
        header_group = QGroupBox()
        header_layout = QHBoxLayout(header_group)
        header_layout.addWidget(QLabel("<h2>📊 EVM Analysis Dashboard</h2>"))
        
        header_layout.addStretch()
        
        header_layout.addWidget(QLabel("<b>Baseline:</b>"))
        self.baseline_selector = QComboBox()
        self.baseline_selector.setMinimumWidth(150)
        header_layout.addWidget(self.baseline_selector)
        
        self.refresh_btn = QPushButton("🔄 Refresh EVM")
        self.refresh_btn.setMinimumHeight(30)
        self.refresh_btn.clicked.connect(self.refresh_data)
        self.refresh_btn.setStyleSheet("background-color: #E3F2FD; font-weight: bold;")
        header_layout.addWidget(self.refresh_btn)
        
        self.export_btn = QPushButton("📥 Export Excel")
        self.export_btn.setMinimumHeight(30)
        self.export_btn.clicked.connect(self.export_to_excel)
        header_layout.addWidget(self.export_btn)
        
        self.help_btn = QPushButton("❓ Help")
        self.help_btn.setMinimumHeight(30)
        self.help_btn.setFixedWidth(80)
        self.help_btn.clicked.connect(self.show_help)
        header_layout.addWidget(self.help_btn)
        
        layout.addWidget(header_group)
        
        # --- Scrollable Content Area ---
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("QScrollArea { border: none; }")
        
        content_widget = QWidget()
        self.content_layout = QVBoxLayout(content_widget)
        
        # 1. Summary Cards Area
        self._setup_summary_cards()
        
        # 2. Executive Summary
        self._setup_health_analysis()
        
        # 3. Charts Area
        self._setup_charts_area()
        
        # 4. Table Area
        self._setup_table_area()
        
        scroll_area.setWidget(content_widget)
        layout.addWidget(scroll_area)
        
        # Initial population of baseline list
        self.update_baseline_list()

    def update_baseline_list(self):
        """Update the baseline dropdown from data manager."""
        current_text = self.baseline_selector.currentText()
        self.baseline_selector.clear()
        baselines = self.main_window.data_manager.get_all_baselines()
        for b in baselines:
            self.baseline_selector.addItem(b.name)
        
        # Restore selection if it still exists
        index = self.baseline_selector.findText(current_text)
        if index >= 0:
            self.baseline_selector.setCurrentIndex(index)
        elif self.baseline_selector.count() > 0:
            self.baseline_selector.setCurrentIndex(0)

    def _setup_summary_cards(self):
        summary_group = QGroupBox("Project-Level Status")
        summary_layout = QHBoxLayout(summary_group)
        self.content_layout.addWidget(summary_group)
        
        self.summary_labels = {}
        # Metrics to display in cards: (Key, Display Label)
        metrics = [
            ("CPI", "Cost Performance Index"),
            ("SPI", "Schedule Performance Index"),
            ("CV", "Cost Variance"),
            ("SV", "Schedule Variance"),
            ("EAC", "Estimate at Completion"),
            ("VAC", "Variance at Completion")
        ]
        
        for key, label in metrics:
            card = QWidget()
            card_v_layout = QVBoxLayout(card)
            
            val_label = QLabel("N/A")
            val_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            val_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50;")
            
            title_label = QLabel(label)
            title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            title_label.setStyleSheet("color: #7f8c8d; font-size: 11px;")
            title_label.setWordWrap(True) # Ensure long text wraps if needed
            
            card_v_layout.addWidget(val_label)
            card_v_layout.addWidget(title_label)
            self.summary_labels[key] = val_label
            summary_layout.addWidget(card)
            
            if key != "VAC": # Add separator
                line = QLabel("|")
                line.setStyleSheet("color: #dcdde1;")
                summary_layout.addWidget(line)

                summary_layout.addWidget(line)

    def _setup_health_analysis(self):
        """Setup the text-based project health summary."""
        group = QGroupBox("Executive Summary")
        layout = QVBoxLayout(group)
        
        self.health_label = QLabel("Select a baseline and click refresh to see analysis.")
        self.health_label.setWordWrap(True)
        self.health_label.setStyleSheet("font-size: 14px; padding: 5px;")
        
        layout.addWidget(self.health_label)
        self.content_layout.addWidget(group)

    def _setup_charts_area(self):
        charts_layout = QHBoxLayout()
        
        # Left Chart: Comparison of Values
        self.val_figure = Figure(figsize=(5, 3), dpi=100)
        self.val_canvas = FigureCanvas(self.val_figure)
        self.val_canvas.setMinimumHeight(250)
        charts_layout.addWidget(self.val_canvas)
        
        # Right Chart: Performance Indices over time (or simple bar for now)
        self.idx_figure = Figure(figsize=(5, 3), dpi=100)
        self.idx_canvas = FigureCanvas(self.idx_figure)
        self.idx_canvas.setMinimumHeight(250)
        charts_layout.addWidget(self.idx_canvas)
        
        self.content_layout.addLayout(charts_layout)

    def _setup_table_area(self):
        table_group = QGroupBox("Task Metric Details")
        table_layout = QVBoxLayout(table_group)
        
        self.table = QTableWidget()
        cols = [
            "WBS", "Task", 
            "Planned Value (PV)", "Earned Value (EV)", "Actual Cost (AC)", "Budget at Completion (BAC)", 
            "Cost Variance (CV)", "Schedule Variance (SV)", 
            "Cost Performance Index (CPI)", "Schedule Performance Index (SPI)"
        ]
        self.table.setColumnCount(len(cols))
        self.table.setHorizontalHeaderLabels(cols)
        
        # Add tooltips to headers
        header_tooltips = {
            "WBS": "Work Breakdown Structure Code",
            "Task": "Task Name",
            "Planned Value (PV)": "Planned Value (BCWS)\nBudgeted cost for work scheduled to be done.",
            "Earned Value (EV)": "Earned Value (BCWP)\nBudgeted cost for work actually performed.",
            "Actual Cost (AC)": "Actual Cost (ACWP)\nActual cost incurred for work performed.",
            "Budget at Completion (BAC)": "Budget at Completion\nTotal planned cost for the task.",
            "Cost Variance (CV)": "Cost Variance (EV - AC)\nDifference between earned value and actual cost.",
            "Schedule Variance (SV)": "Schedule Variance (EV - PV)\nDifference between earned value and planned value.",
            "Cost Performance Index (CPI)": "Cost Performance Index (EV / AC)\nEfficiency of cost spend (>1 is good).",
            "Schedule Performance Index (SPI)": "Schedule Performance Index (EV / PV)\nEfficiency of time usage (>1 is good)."
        }
        
        for i, col_name in enumerate(cols):
            item = self.table.horizontalHeaderItem(i)
            if item and col_name in header_tooltips:
                item.setToolTip(header_tooltips[col_name])
                
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        # Configure column resizing
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch) # Default to stretch
        
        # Set specific columns to resize to contents for compactness
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents) # WBS
        
        # Indices for columns to compact: PV(2), EV(3), AC(4), CV(6), SV(7)
        compact_cols = [2, 3, 4, 6, 7]
        for col_idx in compact_cols:
             header.setSectionResizeMode(col_idx, QHeaderView.ResizeMode.ResizeToContents)
        
        table_layout.addWidget(self.table)
        self.content_layout.addWidget(table_group)

    def clear_view(self):
        """Reset the EVM tab to its initial state."""
        self.baseline_selector.clear()
        
        # Reset labels
        if hasattr(self, 'summary_labels'):
            for label in self.summary_labels.values():
                label.setText("N/A")
                label.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50;")
            
        if hasattr(self, 'health_label'):
            self.health_label.setText("Select a baseline and click refresh to see analysis.")
        
        # Clear charts
        if hasattr(self, 'val_figure'):
            self.val_figure.clear()
            self.val_canvas.draw()
        if hasattr(self, 'idx_figure'):
            self.idx_figure.clear()
            self.idx_canvas.draw()
        
        # Clear table
        if hasattr(self, 'table'):
            self.table.setRowCount(0)

    def show_help(self):
        """Show the EVM help dialog."""
        dialog = EVMHelpDialog(self)
        dialog.exec()

    def update_data_manager(self, data_manager):
        """Update the data manager reference when a new project is loaded."""
        self.calculator.data_manager = data_manager

    def refresh_data(self, silent=False):
        """Recalculate and update the UI."""
        self.update_baseline_list()
        baseline_name = self.baseline_selector.currentText()
        
        if not baseline_name:
            # Check if there are any baselines at all
            if not silent and not self.main_window.data_manager.get_all_baselines():
                QMessageBox.warning(self, "No Baseline", 
                    "No baselines found. Please create a baseline (Project -> Manage Baselines) first to use EVM Analysis.")
            return

        # Calculate metrics
        # We use today as the status date
        df = self.calculator.calculate_metrics(baseline_name)
        
        if df.empty:
            if not silent:
                QMessageBox.information(self, "No Data", "No task data found for the selected baseline.")
            return
            
        summary = self.calculator.get_summary_metrics(df)
        
        # Update UI components
        self.update_summary_display(summary)
        self.update_health_display(summary)
        self.update_charts_display(summary)
        self.update_table_display(df)

    def update_health_display(self, summary):
        """Generate and display a text summary of project health."""
        cpi = summary['CPI']
        spi = summary['SPI']
        eac = summary['EAC']
        bac = summary['BAC']
        symbol = self.main_window.data_manager.settings.currency.symbol
        
        # Determine Cost Status
        if cpi >= 1.05:
            cost_status = "<b style='color: #27ae60;'>Under Budget</b> (Efficient)"
        elif 0.95 <= cpi < 1.05:
            cost_status = "<b style='color: #f39c12;'>On Budget</b>"
        else:
            cost_status = "<b style='color: #e74c3c;'>Over Budget</b> (Inefficient)"
            
        # Determine Schedule Status
        if spi >= 1.05:
            sched_status = "<b style='color: #27ae60;'>Ahead of Schedule</b>"
        elif 0.95 <= spi < 1.05:
            sched_status = "<b style='color: #f39c12;'>On Schedule</b>"
        else:
            sched_status = "<b style='color: #e74c3c;'>Behind Schedule</b>"
            
        # Efficiency Insight
        if cpi < 1.0:
             cost_efficiency = f"For every 1.00 unit of value, we are spending {1/cpi:.2f} units."
        elif cpi > 1.0:
             cost_efficiency = f"For every 1.00 unit spent, we are earning {cpi:.2f} units of value."
        else:
             cost_efficiency = "Cost efficiency is exactly as planned."

        if spi < 1.0:
            sched_efficiency = f"Work is being completed at {spi*100:.1f}% of the planned rate."
        elif spi > 1.0:
            sched_efficiency = f"Work is being completed at {spi*100:.1f}% of the planned rate."
        else:
            sched_efficiency = "Work is progressing exactly as planned."

        text = (
            f"The project is currently {sched_status} (SPI: {spi:.2f}) and {cost_status} (CPI: {cpi:.2f}).<br>"
            f"<b>Performance:</b> {sched_efficiency} {cost_efficiency}<br>"
            f"At this pace, the estimated cost at completion is <b>{symbol}{eac:,.2f}</b> "
            f"(Original Budget: <b>{symbol}{bac:,.2f}</b>)."
        )
        self.health_label.setText(text)

    def update_summary_display(self, summary):
        symbol = self.main_window.data_manager.settings.currency.symbol
        
        # Mapping functions for color coding
        def get_color(val, is_index=True):
            if is_index:
                return "#27ae60" if val >= 1.0 else "#e74c3c"
            return "#27ae60" if val >= 0 else "#e74c3c"

        self.summary_labels["CPI"].setText(f"{summary['CPI']:.2f}")
        self.summary_labels["CPI"].setStyleSheet(f"font-size: 18px; font-weight: bold; color: {get_color(summary['CPI'])};")
        
        self.summary_labels["SPI"].setText(f"{summary['SPI']:.2f}")
        self.summary_labels["SPI"].setStyleSheet(f"font-size: 18px; font-weight: bold; color: {get_color(summary['SPI'])};")
        
        self.summary_labels["CV"].setText(f"{symbol}{summary['CV']:,.2f}")
        self.summary_labels["CV"].setStyleSheet(f"font-size: 18px; font-weight: bold; color: {get_color(summary['CV'], False)};")
        
        self.summary_labels["SV"].setText(f"{symbol}{summary['SV']:,.2f}")
        self.summary_labels["SV"].setStyleSheet(f"font-size: 18px; font-weight: bold; color: {get_color(summary['SV'], False)};")
        
        self.summary_labels["EAC"].setText(f"{symbol}{summary['EAC']:,.2f}")
        self.summary_labels["VAC"].setText(f"{symbol}{summary['VAC']:,.2f}")
        self.summary_labels["VAC"].setStyleSheet(f"font-size: 18px; font-weight: bold; color: {get_color(summary['VAC'], False)};")

    def update_charts_display(self, summary):
        # 1. Budget Comparison Chart
        self.val_figure.clear()
        ax1 = self.val_figure.add_subplot(111)
        labels = ['PV', 'EV', 'AC', 'BAC']
        values = [summary['PV'], summary['EV'], summary['AC'], summary['BAC']]
        colors = ['#3498db', '#2ecc71', '#e74c3c', '#f1c40f']
        
        bars = ax1.bar(labels, values, color=colors, alpha=0.8)
        ax1.set_title("EVM Value Comparison", fontsize=11, fontweight='bold')
        ax1.grid(axis='y', linestyle='--', alpha=0.6)
        
        # Add values on top of bars
        symbol = self.main_window.data_manager.settings.currency.symbol
        for bar in bars:
            height = bar.get_height()
            ax1.text(bar.get_x() + bar.get_width()/2., height + (max(values)*0.01),
                    f'{symbol}{height:,.0f}', ha='center', va='bottom', fontsize=8)
        
        self.val_canvas.draw()
        
        # 2. Performance Indices Chart
        self.idx_figure.clear()
        ax2 = self.idx_figure.add_subplot(111)
        indices = ['CPI', 'SPI']
        idx_vals = [summary['CPI'], summary['SPI']]
        idx_colors = ['#27ae60' if v >= 1.0 else '#e74c3c' for v in idx_vals]
        
        ax2.bar(indices, idx_vals, color=idx_colors, width=0.5)
        ax2.axhline(y=1.0, color='#2c3e50', linestyle='--', alpha=0.6, label='Target (1.0)')
        ax2.set_ylim(0, max(1.5, max(idx_vals) * 1.2))
        ax2.set_title("Performance Indices", fontsize=11, fontweight='bold')
        ax2.legend(fontsize=8)
        
        for i, v in enumerate(idx_vals):
            ax2.text(i, v + 0.05, f'{v:.2f}', ha='center', va='bottom', fontweight='bold')
            
        self.idx_canvas.draw()

    def update_table_display(self, df):
        self.table.setRowCount(len(df))
        symbol = self.main_window.data_manager.settings.currency.symbol
        
        for i, row in df.iterrows():
            self.table.setItem(i, 0, QTableWidgetItem(str(row['WBS'])))
            self.table.setItem(i, 1, QTableWidgetItem(str(row['Task'])))
            
            # Formatted financials
            for j, key in enumerate(['PV', 'EV', 'AC', 'BAC', 'CV', 'SV']):
                item = QTableWidgetItem(f"{symbol}{row[key]:,.2f}")
                if key in ['CV', 'SV']:
                    color = Qt.GlobalColor.darkGreen if row[key] >= 0 else Qt.GlobalColor.red
                    item.setForeground(color)
                self.table.setItem(i, j + 2, item)
            
            # Indices
            self.table.setItem(i, 8, QTableWidgetItem(f"{row['CPI']:.2f}"))
            self.table.setItem(i, 9, QTableWidgetItem(f"{row['SPI']:.2f}"))
            
            # Apply summary task formatting
            if row['is_summary']:
                for col in range(self.table.columnCount()):
                    font = self.table.item(i, col).font()
                    font.setBold(True)
                    self.table.item(i, col).setFont(font)
                    self.table.item(i, col).setBackground(Qt.GlobalColor.lightGray)

    def export_to_excel(self):
        """Export the EVM data to an Excel file."""
        baseline_name = self.baseline_selector.currentText()
        if not baseline_name:
            return
            
        df = self.calculator.calculate_metrics(baseline_name)
        if df.empty:
            return
            
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export EVM Analysis", f"EVM_Analysis_{baseline_name}.xlsx", "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                summary = self.calculator.get_summary_metrics(df)
                
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    # 1. Executive Summary Sheet
                    summary_data = {
                        'Metric': [
                            'Cost Performance Index (CPI)', 'Schedule Performance Index (SPI)',
                            'Cost Variance (CV)', 'Schedule Variance (SV)',
                            'Planned Value (PV)', 'Earned Value (EV)', 'Actual Cost (AC)',
                            'Budget at Completion (BAC)', 'Estimate at Completion (EAC)', 'Variance at Completion (VAC)'
                        ],
                        'Value': [
                            round(summary['CPI'], 2), round(summary['SPI'], 2),
                            round(summary['CV'], 2), round(summary['SV'], 2),
                            round(summary['PV'], 2), round(summary['EV'], 2), round(summary['AC'], 2),
                            round(summary['BAC'], 2), round(summary['EAC'], 2), round(summary['VAC'], 2)
                        ],
                        'Description': [
                            'Efficiency of cost spend (>1.0 is Under Budget)', 
                            'Efficiency of time usage (>1.0 is Ahead of Schedule)',
                            'Difference between Value Earned and Cost Spent (Positive is Good)', 
                            'Difference between Value Earned and Value Planned (Positive is Good)',
                            'Budgeted cost for work scheduled so far', 
                            'Budgeted cost for work actually performed', 
                            'Actual cost incurred',
                            'Total planned cost of the project', 
                            'Predicted total cost based on current performance', 
                            'Projected budget surplus (+) or deficit (-)'
                        ]
                    }
                    pd.DataFrame(summary_data).to_excel(writer, sheet_name='Executive Summary', index=False)
                    
                    # 2. Detailed Metrics Sheet
                    # Rename columns for clarity in export
                    export_df = df.drop(columns=['is_summary', 'Task ID']).rename(columns={
                        'PV': 'Planned Value (PV)',
                        'EV': 'Earned Value (EV)',
                        'AC': 'Actual Cost (AC)',
                        'BAC': 'Budget at Completion (BAC)',
                        'CV': 'Cost Variance (CV)',
                        'SV': 'Schedule Variance (SV)',
                        'EAC': 'Estimate at Completion (EAC)',
                        'VAC': 'Variance at Completion (VAC)',
                        'CPI': 'Cost Performance Index (CPI)',
                        'SPI': 'Schedule Performance Index (SPI)'
                    })
                    export_df.to_excel(writer, sheet_name='Task Metrics', index=False)
                    
                    # 3. Glossary Sheet
                    help_data = {
                        'Term': ['PV', 'EV', 'AC', 'BAC', 'CV', 'SV', 'CPI', 'SPI', 'EAC', 'VAC'],
                        'Full Name': [
                            'Planned Value', 'Earned Value', 'Actual Cost', 'Budget at Completion',
                            'Cost Variance', 'Schedule Variance', 'Cost Performance Index', 'Schedule Performance Index',
                            'Estimate at Completion', 'Variance at Completion'
                        ],
                        'Definition': [
                            'The budgeted cost for the work scheduled to be completed by the status date.',
                            'The budgeted cost for the work actually performed to date.',
                            'The actual cost incurred for the work performed so far.',
                            'The total planned cost of the project/task when completed.',
                            'EV - AC. Positive value indicates under budget.',
                            'EV - PV. Positive value indicates ahead of schedule.',
                            'EV / AC. Measure of cost efficiency. 1.0 is on budget.',
                            'EV / PV. Measure of schedule efficiency. 1.0 is on schedule.',
                            'BAC / CPI. Forecasted total cost based on current trends.',
                            'BAC - EAC. Projected variance at the end of the project.'
                        ]
                    }
                    pd.DataFrame(help_data).to_excel(writer, sheet_name='Glossary', index=False)

                # Format columns auto-width (using openpyxl)
                from openpyxl import load_workbook
                wb = load_workbook(file_path)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        ws.column_dimensions[column_letter].width = adjusted_width
                wb.save(file_path)

                msg_box = QMessageBox(self)
                msg_box.setWindowTitle("Success")
                msg_box.setText(f"EVM analysis exported successfully with detailed sheets to:\n{file_path}")
                msg_box.setIcon(QMessageBox.Icon.Information)
                
                open_button = msg_box.addButton("Open File", QMessageBox.ButtonRole.ActionRole)
                msg_box.addButton(QMessageBox.StandardButton.Ok)
                
                msg_box.exec()
                
                if msg_box.clickedButton() == open_button:
                    import os
                    os.startfile(file_path)
            except Exception as e:
                QMessageBox.critical(self, "Export Error", f"Failed to export Excel:\n{str(e)}")
