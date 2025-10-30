"""
Exporter - Handles Excel and JSON import/export
Enhanced with hierarchy, dependency types, project name, and backward compatibility
"""

import pandas as pd
import json
from typing import Dict, Any, Tuple
from datetime import datetime
from data_manager import DataManager, Task, Resource, DependencyType
from calendar_manager import CalendarManager

class Exporter:
    """Handles data import/export operations with enhanced features"""
    
    @staticmethod
    def export_to_excel(data_manager: DataManager, filepath: str) -> bool:
        """Export project to Excel file with hierarchy and dependency types"""
        try:
            # Get duration unit label
            duration_label = data_manager.settings.get_duration_label()
            
            # Prepare tasks dataframe with hierarchy
            tasks_data = []
            
            # Helper function to add tasks in hierarchical order
            def add_task_and_children(task: Task, level: int = 0):
                # Calculate status
                status = task.get_status_text()
                status_color = task.get_status_color()
                
                # Format predecessors with dependency types
                pred_text = ', '.join([
                    f"{pred_id} ({DependencyType[dep_type].value})" 
                    for pred_id, dep_type in task.predecessors
                ])
                
                # Indent task name based on level
                indent = "  " * level
                display_name = f"{indent}{task.name}"
                # *** ADD MILESTONE INDICATOR ***
                if task.is_milestone:
                    display_name = f"{indent}◆ {task.name}"
                elif task.is_summary:
                    display_name = f"{indent}▶ {task.name}"
                
                duration = task.get_duration(
                    data_manager.settings.duration_unit,
                    data_manager.calendar_manager
                )
                
                tasks_data.append({
                    'Task ID': task.id,
                    'Task Name': display_name,
                    'Type': 'Milestone' if task.is_milestone else ('Summary' if task.is_summary else 'Task'),  # *** ADD TYPE COLUMN ***
                    'Level': level,
                    'Parent ID': task.parent_id if task.parent_id else '',
                    'Is Summary': 'Yes' if task.is_summary else 'No',
                    'Is Milestone': 'Yes' if task.is_milestone else 'No',  # *** ADD THIS ***
                    'Start Date': task.start_date.strftime('%Y-%m-%d'),
                    'End Date': task.end_date.strftime('%Y-%m-%d') if not task.is_milestone else 'Milestone',
                    duration_label: 0 if task.is_milestone else (f"{duration:.1f}" if data_manager.settings.duration_unit == DurationUnit.HOURS else int(duration)),
                    '% Complete': task.percent_complete,
                    'Status': status,
                    'Status Color': status_color,
                    'Predecessors': pred_text,
                    'Assigned Resources': ','.join(task.assigned_resources),
                    'Notes': task.notes
                })
                
                # Add children
                children = data_manager.get_child_tasks(task.id)
                for child in sorted(children, key=lambda t: t.start_date):
                    add_task_and_children(child, level + 1)
            
            # Add all top-level tasks
            top_level = data_manager.get_top_level_tasks()
            for task in sorted(top_level, key=lambda t: t.start_date):
                add_task_and_children(task, 0)
            
            # Prepare resources dataframe
            resources_data = []
            allocation = data_manager.get_resource_allocation()
            for resource in data_manager.get_all_resources():
                res_alloc = allocation.get(resource.name, {})
                resources_data.append({
                    'Resource Name': resource.name,
                    'Max Hours/Day': resource.max_hours_per_day,
                    'Total Hours Allocated': res_alloc.get('total_hours', 0),
                    'Tasks Assigned': res_alloc.get('tasks_assigned', 0),
                    'Exceptions': ','.join(resource.exceptions)
                })
            
            # Prepare summary dataframe
            summary_data = {
                'Metric': [
                    'Project Name',
                    'Project Start Date',
                    'Project End Date',
                    'Total Tasks',
                    'Summary Tasks',
                    'Work Tasks',
                    'Total Resources',
                    'Overall % Complete',
                    'Total Project Duration (days)',
                    'Export Date',
                    'File Version'
                ],
                'Value': [
                    data_manager.project_name,
                    data_manager.get_project_start_date().strftime('%Y-%m-%d') 
                        if data_manager.get_project_start_date() else 'N/A',
                    data_manager.get_project_end_date().strftime('%Y-%m-%d')
                        if data_manager.get_project_end_date() else 'N/A',
                    len(data_manager.tasks),
                    sum(1 for t in data_manager.tasks if t.is_summary),
                    sum(1 for t in data_manager.tasks if not t.is_summary),
                    len(data_manager.resources),
                    f"{data_manager.get_overall_completion():.1f}%",
                    (data_manager.get_project_end_date() - data_manager.get_project_start_date()).days + 1
                        if data_manager.get_project_start_date() and data_manager.get_project_end_date() else 'N/A',
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    '2.0'
                ]
            }
            
            # Prepare status breakdown
            status_counts = {}
            for task in data_manager.tasks:
                if not task.is_summary:
                    status = task.get_status_text()
                    status_counts[status] = status_counts.get(status, 0) + 1
            
            status_data = {
                'Status': list(status_counts.keys()),
                'Count': list(status_counts.values())
            }
            
            # Prepare dependency types breakdown
            dep_type_counts = {}
            for task in data_manager.tasks:
                for _, dep_type in task.predecessors:
                    try:
                        dep_name = DependencyType[dep_type].value
                        dep_type_counts[dep_name] = dep_type_counts.get(dep_name, 0) + 1
                    except:
                        pass
            
            dep_data = {
                'Dependency Type': list(dep_type_counts.keys()) if dep_type_counts else ['None'],
                'Count': list(dep_type_counts.values()) if dep_type_counts else [0]
            }
            
            # Create Excel writer with formatting
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Write summary first
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                
                # Format summary sheet
                workbook = writer.book
                summary_sheet = writer.sheets['Summary']
                summary_sheet.column_dimensions['A'].width = 30
                summary_sheet.column_dimensions['B'].width = 25
                
                # Write tasks
                if tasks_data:
                    df_tasks = pd.DataFrame(tasks_data)
                    df_tasks.to_excel(writer, sheet_name='Tasks', index=False)
                    
                    # Format tasks sheet
                    task_sheet = writer.sheets['Tasks']
                    task_sheet.column_dimensions['B'].width = 40  # Task Name
                    task_sheet.column_dimensions['L'].width = 30  # Predecessors
                    task_sheet.column_dimensions['M'].width = 25  # Resources
                    
                    # Apply color coding to status column
                    from openpyxl.styles import PatternFill
                    
                    status_col_idx = df_tasks.columns.get_loc('Status Color') + 1
                    for row_idx, row in df_tasks.iterrows():
                        cell = task_sheet.cell(row=row_idx + 2, column=status_col_idx)
                        color_name = row['Status Color']
                        
                        color_map = {
                            'red': 'FFCDD2',
                            'green': 'C8E6C9',
                            'grey': 'F5F5F5',
                            'blue': 'BBDEFB'
                        }
                        
                        if color_name in color_map:
                            fill = PatternFill(start_color=color_map[color_name],
                                             end_color=color_map[color_name],
                                             fill_type='solid')
                            # Apply to entire row
                            for col in range(1, len(df_tasks.columns) + 1):
                                task_sheet.cell(row=row_idx + 2, column=col).fill = fill
                
                # Write resources
                if resources_data:
                    df_resources = pd.DataFrame(resources_data)
                    df_resources.to_excel(writer, sheet_name='Resources', index=False)
                    
                    resource_sheet = writer.sheets['Resources']
                    resource_sheet.column_dimensions['A'].width = 20
                
                # Write status breakdown
                df_status = pd.DataFrame(status_data)
                df_status.to_excel(writer, sheet_name='Status Breakdown', index=False)
                
                # Write dependency breakdown
                df_deps = pd.DataFrame(dep_data)
                df_deps.to_excel(writer, sheet_name='Dependencies', index=False)
            
            return True
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    @staticmethod
    def import_from_excel(filepath: str) -> Tuple[DataManager, bool]:
        """Import project from Excel file with backward compatibility"""
        data_manager = DataManager()
        
        try:
            # Try to read summary for project name
            try:
                df_summary = pd.read_excel(filepath, sheet_name='Summary')
                for _, row in df_summary.iterrows():
                    if row['Metric'] == 'Project Name':
                        data_manager.project_name = str(row['Value'])
                        break
            except:
                pass
            
            # Read tasks
            df_tasks = pd.read_excel(filepath, sheet_name='Tasks')
            
            # Check if this is new format (with Level and Parent ID columns)
            has_hierarchy = 'Level' in df_tasks.columns and 'Parent ID' in df_tasks.columns
            has_is_summary = 'Is Summary' in df_tasks.columns
            
            for _, row in df_tasks.iterrows():
                # Parse predecessors with dependency types
                predecessors = []
                if pd.notna(row.get('Predecessors', '')) and str(row['Predecessors']).strip():
                    pred_str = str(row['Predecessors'])
                    
                    # Check if new format with dependency types
                    if '(' in pred_str:
                        # New format: "1 (Finish-to-Start), 2 (Start-to-Start)"
                        for pred_item in pred_str.split(','):
                            pred_item = pred_item.strip()
                            if '(' in pred_item:
                                # Extract ID and type
                                parts = pred_item.split('(')
                                pred_id = int(parts[0].strip())
                                dep_type_text = parts[1].replace(')', '').strip()
                                
                                # Convert text to enum name
                                dep_type = None
                                for dt in DependencyType:
                                    if dt.value == dep_type_text:
                                        dep_type = dt.name
                                        break
                                
                                if dep_type is None:
                                    dep_type = DependencyType.FS.name
                                
                                predecessors.append((pred_id, dep_type))
                            else:
                                # Fallback: just ID, assume FS
                                try:
                                    pred_id = int(pred_item.strip())
                                    predecessors.append((pred_id, DependencyType.FS.name))
                                except:
                                    pass
                    else:
                        # Old format: just IDs "1,2,3"
                        for pred_id_str in pred_str.split(','):
                            try:
                                pred_id = int(pred_id_str.strip())
                                predecessors.append((pred_id, DependencyType.FS.name))
                            except:
                                pass
                
                # Parse resources
                assigned_resources = []
                if pd.notna(row.get('Assigned Resources', '')) and str(row['Assigned Resources']).strip():
                    assigned_resources = [x.strip() for x in str(row['Assigned Resources']).split(',')]
                
                # Get parent ID
                parent_id = None
                if has_hierarchy and pd.notna(row.get('Parent ID', '')):
                    try:
                        parent_id = int(row['Parent ID'])
                    except:
                        pass
                
                # Get is_summary flag
                is_summary = False
                if has_is_summary and pd.notna(row.get('Is Summary', '')):
                    is_summary = str(row['Is Summary']).lower() in ['yes', 'true', '1']
                
                # Clean task name (remove indentation and summary marker)
                task_name = str(row['Task Name']).strip()
                task_name = task_name.lstrip(' ▶')
                
                                # Get is_milestone flag
                is_milestone = False
                if has_is_milestone and pd.notna(row.get('Is Milestone', '')):
                    is_milestone = str(row['Is Milestone']).lower() in ['yes', 'true', '1']
                
                task = Task(
                    name=task_name,
                    start_date=pd.to_datetime(row['Start Date']).to_pydatetime(),
                    end_date=pd.to_datetime(row['End Date']).to_pydatetime() if not is_milestone else pd.to_datetime(row['Start Date']).to_pydatetime(),
                    percent_complete=int(row.get('% Complete', 0)),
                    predecessors=predecessors,
                    assigned_resources=assigned_resources,
                    notes=str(row.get('Notes', '')),
                    task_id=int(row['Task ID']),
                    parent_id=parent_id,
                    is_summary=is_summary,
                    is_milestone=is_milestone  # *** ADD THIS ***
                )
                data_manager.tasks.append(task)
            
            # Read resources if sheet exists
            try:
                df_resources = pd.read_excel(filepath, sheet_name='Resources')
                
                for _, row in df_resources.iterrows():
                    exceptions = []
                    if pd.notna(row.get('Exceptions', '')) and str(row['Exceptions']).strip():
                        exceptions = [x.strip() for x in str(row['Exceptions']).split(',')]
                    
                    resource = Resource(
                        name=str(row['Resource Name']),
                        max_hours_per_day=float(row.get('Max Hours/Day', 8.0)),
                        exceptions=exceptions
                    )
                    data_manager.resources.append(resource)
            except:
                pass
            
            return data_manager, True
            
        except Exception as e:
            print(f"Error importing from Excel: {e}")
            import traceback
            traceback.print_exc()
            return data_manager, False
    
    @staticmethod
    def save_to_json(data_manager: DataManager, calendar_manager: CalendarManager, 
                     filepath: str) -> bool:
        """Save project to JSON file with all enhanced features"""
        try:
            data = {
                'version': '2.1',
                'project_name': data_manager.project_name,
                'project_data': data_manager.to_dict(),
                'calendar_settings': calendar_manager.to_dict(),
                'saved_at': datetime.now().isoformat(),
                'metadata': {
                    'total_tasks': len(data_manager.tasks),
                    'summary_tasks': sum(1 for t in data_manager.tasks if t.is_summary),
                    'work_tasks': sum(1 for t in data_manager.tasks if not t.is_summary),
                    'total_resources': len(data_manager.resources),
                    'overall_completion': data_manager.get_overall_completion(),
                    'duration_unit': data_manager.settings.duration_unit.value
                }
            }
            
            with open(filepath, 'w') as f:
                json.dump(data, f, indent=2)
            
            return True
            
        except Exception as e:
            print(f"Error saving to JSON: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    @staticmethod
    def load_from_json(filepath: str) -> Tuple[DataManager, CalendarManager, bool]:
        """Load project from JSON file with backward compatibility"""
        data_manager = DataManager()
        calendar_manager = CalendarManager()
        
        try:
            with open(filepath, 'r') as f:
                data = json.load(f)
            
            # Check version for backward compatibility
            version = data.get('version', '1.0')
            
            # Load project name (new in v2.0)
            if 'project_name' in data:
                data_manager.project_name = data['project_name']
            elif 'project_data' in data and 'project_name' in data['project_data']:
                data_manager.project_name = data['project_data']['project_name']
            else:
                # Try to extract from filename
                import os
                filename = os.path.basename(filepath)
                data_manager.project_name = os.path.splitext(filename)[0].replace('_', ' ')
            
            # Load project data
            data_manager.from_dict(data.get('project_data', {}))
            
            # Load calendar settings
            calendar_manager.from_dict(data.get('calendar_settings', {}))
            
            # Set calendar manager reference
            data_manager.calendar_manager = calendar_manager
            
            return data_manager, calendar_manager, True
            
        except Exception as e:
            print(f"Error loading from JSON: {e}")
            import traceback
            traceback.print_exc()
            return data_manager, calendar_manager, False
    
    @staticmethod
    def export_to_csv(data_manager: DataManager, filepath: str) -> bool:
        """Export tasks to CSV format (flat structure)"""
        try:
            tasks_data = []
            
            for task in sorted(data_manager.get_all_tasks(), key=lambda t: t.id):
                # Format predecessors
                pred_text = ', '.join([
                    f"{pred_id} ({DependencyType[dep_type].value})" 
                    for pred_id, dep_type in task.predecessors
                ])
                
                tasks_data.append({
                    'ID': task.id,
                    'Name': task.name,
                    'Parent_ID': task.parent_id if task.parent_id else '',
                    'Is_Summary': task.is_summary,
                    'Start': task.start_date.strftime('%Y-%m-%d'),
                    'End': task.end_date.strftime('%Y-%m-%d'),
                    'Duration': task.duration,
                    'Complete': task.percent_complete,
                    'Status': task.get_status_text(),
                    'Predecessors': pred_text,
                    'Resources': ','.join(task.assigned_resources),
                    'Notes': task.notes
                })
            
            df = pd.DataFrame(tasks_data)
            df.to_csv(filepath, index=False)
            
            return True
            
        except Exception as e:
            print(f"Error exporting to CSV: {e}")
            return False
    
    @staticmethod
    def export_project_report(data_manager: DataManager, filepath: str) -> bool:
        """Export comprehensive project report to Excel"""
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Project Overview
                overview_data = {
                    'Item': [
                        'Project Name',
                        'Start Date',
                        'End Date',
                        'Duration (days)',
                        'Overall Completion (%)',
                        'Total Tasks',
                        'Summary Tasks',
                        'Work Tasks',
                        'Resources',
                        'Report Date'
                    ],
                    'Value': [
                        data_manager.project_name,
                        data_manager.get_project_start_date().strftime('%Y-%m-%d') 
                            if data_manager.get_project_start_date() else 'N/A',
                        data_manager.get_project_end_date().strftime('%Y-%m-%d')
                            if data_manager.get_project_end_date() else 'N/A',
                        (data_manager.get_project_end_date() - data_manager.get_project_start_date()).days + 1
                            if data_manager.get_project_start_date() and data_manager.get_project_end_date() else 'N/A',
                        f"{data_manager.get_overall_completion():.1f}",
                        len(data_manager.tasks),
                        sum(1 for t in data_manager.tasks if t.is_summary),
                        sum(1 for t in data_manager.tasks if not t.is_summary),
                        len(data_manager.resources),
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    ]
                }
                pd.DataFrame(overview_data).to_excel(writer, sheet_name='Overview', index=False)
                
                # Task Status Summary
                status_summary = {}
                for task in data_manager.tasks:
                    if not task.is_summary:
                        status = task.get_status_text()
                        status_summary[status] = status_summary.get(status, 0) + 1
                
                status_data = {
                    'Status': list(status_summary.keys()),
                    'Count': list(status_summary.values()),
                    'Percentage': [f"{(count/sum(status_summary.values())*100):.1f}" 
                                  for count in status_summary.values()]
                }
                pd.DataFrame(status_data).to_excel(writer, sheet_name='Status Summary', index=False)
                
                # Critical Tasks (Overdue)
                critical_tasks = []
                for task in data_manager.tasks:
                    if not task.is_summary and task.get_status_text() == 'Overdue':
                        critical_tasks.append({
                            'ID': task.id,
                            'Task': task.name,
                            'End Date': task.end_date.strftime('%Y-%m-%d'),
                            'Days Overdue': (datetime.now() - task.end_date).days,
                            'Complete': f"{task.percent_complete}%",
                            'Resources': ', '.join(task.assigned_resources)
                        })
                
                if critical_tasks:
                    pd.DataFrame(critical_tasks).to_excel(writer, sheet_name='Critical Tasks', index=False)
                
                # Resource Utilization
                allocation = data_manager.get_resource_allocation()
                resource_data = []
                for resource in data_manager.resources:
                    alloc = allocation.get(resource.name, {})
                    resource_data.append({
                        'Resource': resource.name,
                        'Max Hours/Day': resource.max_hours_per_day,
                        'Total Hours': f"{alloc.get('total_hours', 0):.1f}",
                        'Tasks': alloc.get('tasks_assigned', 0),
                        'Utilization': f"{(alloc.get('total_hours', 0) / (resource.max_hours_per_day * 260)):.1%}"
                    })
                
                pd.DataFrame(resource_data).to_excel(writer, sheet_name='Resources', index=False)
                
                # Dependency Analysis
                dep_analysis = {
                    'Finish-to-Start': 0,
                    'Start-to-Start': 0,
                    'Finish-to-Finish': 0,
                    'Start-to-Finish': 0
                }
                
                for task in data_manager.tasks:
                    for _, dep_type in task.predecessors:
                        try:
                            dep_name = DependencyType[dep_type].value
                            dep_analysis[dep_name] = dep_analysis.get(dep_name, 0) + 1
                        except:
                            pass
                
                dep_data = {
                    'Dependency Type': list(dep_analysis.keys()),
                    'Count': list(dep_analysis.values())
                }
                pd.DataFrame(dep_data).to_excel(writer, sheet_name='Dependencies', index=False)
            
            return True
            
        except Exception as e:
            print(f"Error exporting project report: {e}")
            import traceback
            traceback.print_exc()
            return False