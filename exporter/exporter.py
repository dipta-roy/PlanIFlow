"""
Exporter - Handles Excel and JSON import/export
"""
import json
import re
from typing import Tuple
from datetime import datetime
import pandas as pd
from data_manager.manager import DataManager
from data_manager.models import Task, Resource, DependencyType, ScheduleType
from settings_manager.settings_manager import DateFormat, DurationUnit
from calendar_manager.calendar_manager import CalendarManager

class Exporter:
    """Handles data import/export operations with enhanced features"""
    
    @staticmethod
    def _get_date_format_string(date_format: DateFormat) -> str:
        if date_format == DateFormat.DD_MM_YYYY:
            return '%d-%m-%Y %H:%M:%S'
        elif date_format == DateFormat.DD_MMM_YYYY:
            return '%d-%b-%Y %H:%M:%S'
        else: # DateFormat.YYYY_MM_DD
            return '%Y-%m-%dT%H:%M:%S'

    @staticmethod
    def export_to_excel(data_manager: DataManager, filepath: str) -> bool:
        """Export project to Excel file with hierarchy and dependency types"""
        try:
            # Get duration unit label
            duration_label = data_manager.settings.get_duration_label()
            date_format_enum = data_manager.settings.default_date_format
            date_format_str = Exporter._get_date_format_string(date_format_enum)
            
            # Prepare tasks dataframe with hierarchy
            tasks_data = []
            
            # Helper function to add tasks in hierarchical order
            def add_task_and_children(task: Task, level: int = 0):
                # Calculate status
                status = task.get_status_text()
                status_color = task.get_status_color()
                
                pred_text = ', '.join([
                    f"{pred_id} ({DependencyType[dep_type].value}{f'{lag_days:+d}d' if lag_days != 0 else ''})" 
                    for pred_id, dep_type, lag_days in task.predecessors
                ])

                # Format assigned resources with allocation
                resources_text = ', '.join([f"{name} ({alloc}%)" for name, alloc in task.assigned_resources])
                
                # Task name without indentation
                display_name = task.name
                
                duration = task.get_duration(
                    data_manager.settings.duration_unit,
                    data_manager.calendar_manager
                )
                
                tasks_data.append({
                    'Task ID': task.id,
                    'WBS': task.wbs,
                    'Task Name': display_name,
                    'Type': 'Milestone' if task.is_milestone else ('Summary' if task.is_summary else 'Task'),  # *** ADD TYPE COLUMN ***
                    'Schedule Mode': task.schedule_type.value,
                    'Level': level,
                    'Parent ID': task.parent_id if task.parent_id else '',
                    'Is Summary': 'Yes' if task.is_summary else 'No',
                    'Is Milestone': 'Yes' if task.is_milestone else 'No',  # *** ADD THIS ***
                    'Start Date': task.start_date.strftime(date_format_str),
                    'End Date': task.end_date.strftime(date_format_str) if not task.is_milestone else task.start_date.strftime(date_format_str),
                    duration_label: 0 if task.is_milestone else (f"{duration:.1f}" if data_manager.settings.duration_unit == DurationUnit.HOURS else int(duration)),
                    '% Complete': task.percent_complete,
                    'Status': status,
                    'Status Color': status_color,
                    'Predecessors': pred_text,
                    'Assigned Resources': ', '.join([f"{name} ({alloc}%)" for name, alloc in task.assigned_resources]),
                    'Notes': task.notes,
                    # Font styling properties
                    'Font Family': getattr(task, 'font_family', 'Arial'),
                    'Font Size': getattr(task, 'font_size', 10),
                    'Font Color': getattr(task, 'font_color', '#000000'),
                    'Background Color': getattr(task, 'background_color', '#FFFFFF'),
                    'Font Bold': 'Yes' if getattr(task, 'font_bold', False) else 'No',
                    'Font Italic': 'Yes' if getattr(task, 'font_italic', False) else 'No',
                    'Font Underline': 'Yes' if getattr(task, 'font_underline', False) else 'No'
                })
                
                # Add children
                children = data_manager.get_child_tasks(task.id)
                for child in sorted(children, key=lambda t: t.start_date):
                    add_task_and_children(child, level + 1)
            
            # Add all top-level tasks
            top_level = data_manager.get_top_level_tasks()
            for task in sorted(top_level, key=lambda t: t.start_date):
                add_task_and_children(task, 0)
            
            # Prepare resources dataframe
            resources_data = []
            allocation = data_manager.get_resource_allocation()
            for resource in data_manager.get_all_resources():
                res_alloc = allocation.get(resource.name, {})
                resources_data.append({
                    'Resource Name': resource.name,
                    'Max Hours/Day': resource.max_hours_per_day,
                    'Total Hours Allocated': res_alloc.get('total_hours', 0),
                    'Tasks Assigned': res_alloc.get('tasks_assigned', 0),
                    'Exceptions': ','.join(resource.exceptions)
                })
            
            # Prepare summary dataframe
            summary_data = {
                'Metric': [
                    'Project Name',
                    'Project Start Date',
                    'Project End Date',
                    'Working Hours Start',
                    'Working Hours End',
                    'Total Tasks',
                    'Summary Tasks',
                    'Work Tasks',
                    'Total Resources',
                    'Overall % Complete',
                    'Total Project Duration (days)',
                    'Export Date',
                    'File Version',
                    'Date Format'
                ],
                'Value': [
                    data_manager.project_name,
                    data_manager.get_project_start_date().strftime(date_format_str) 
                        if data_manager.get_project_start_date() else 'N/A',
                    data_manager.get_project_end_date().strftime(date_format_str)
                        if data_manager.get_project_end_date() else 'N/A',
                    data_manager.calendar_manager.working_hours_start if data_manager.calendar_manager else '08:00',
                    data_manager.calendar_manager.working_hours_end if data_manager.calendar_manager else '16:00',
                    len(data_manager.tasks),
                    sum(1 for t in data_manager.tasks if t.is_summary),
                    sum(1 for t in data_manager.tasks if not t.is_summary),
                    len(data_manager.resources),
                    f"{data_manager.get_overall_completion():.1f}%",
                    (data_manager.get_project_end_date() - data_manager.get_project_start_date()).days + 1
                        if data_manager.get_project_start_date() and data_manager.get_project_end_date() else 'N/A',
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    '2.3',
                    date_format_enum.value
                ]
            }
            
            # Prepare status breakdown
            status_counts = {}
            for task in data_manager.tasks:
                if not task.is_summary:
                    status = task.get_status_text()
                    status_counts[status] = status_counts.get(status, 0) + 1
            
            status_data = {
                'Status': list(status_counts.keys()),
                'Count': list(status_counts.values())
            }
            
            # Prepare dependency types breakdown
            dep_type_counts = {}
            for task in data_manager.tasks:
                for pred_id, dep_type, lag_days in task.predecessors:
                    try:
                        dep_name = DependencyType[dep_type].value
                        dep_type_counts[dep_name] = dep_type_counts.get(dep_name, 0) + 1
                    except:
                        pass
            
            dep_data = {
                'Dependency Type': list(dep_type_counts.keys()) if dep_type_counts else ['None'],
                'Count': list(dep_type_counts.values()) if dep_type_counts else [0]
            }
            
            # Create Excel writer with formatting
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Write summary first
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                
                # Format summary sheet
                workbook = writer.book
                summary_sheet = writer.sheets['Summary']
                for col in summary_sheet.columns:
                    max_length = max(len(str(cell.value)) for cell in col)
                    summary_sheet.column_dimensions[col[0].column_letter].width = max_length + 2
                
                # Write tasks
                if tasks_data:
                    df_tasks = pd.DataFrame(tasks_data)
                    df_tasks.to_excel(writer, sheet_name='Tasks', index=False)
                    
                    # Format tasks sheet
                    task_sheet = writer.sheets['Tasks']
                    for col in task_sheet.columns:
                        max_length = max(len(str(cell.value)) for cell in col)
                        task_sheet.column_dimensions[col[0].column_letter].width = max_length + 2
                    
                    # Apply color coding to status column and font styling to Task Name
                    from openpyxl.styles import PatternFill, Font
                    
                    status_col_idx = df_tasks.columns.get_loc('Status Color') + 1
                    task_name_col_idx = df_tasks.columns.get_loc('Task Name') + 1
                    
                    for row_idx, row in df_tasks.iterrows():
                        # Apply status color to entire row
                        cell = task_sheet.cell(row=row_idx + 2, column=status_col_idx)
                        color_name = row['Status Color']
                        
                        color_map = {
                            'red': 'FFCDD2',
                            'green': 'C8E6C9',
                            'grey': 'F5F5F5',
                            'blue': 'BBDEFB'
                        }
                        
                        if color_name in color_map:
                            fill = PatternFill(start_color=color_map[color_name],
                                             end_color=color_map[color_name],
                                             fill_type='solid')
                            # Apply to entire row
                            for col in range(1, len(df_tasks.columns) + 1):
                                task_sheet.cell(row=row_idx + 2, column=col).fill = fill
                        
                        # Apply font styling to Task Name cell
                        task_name_cell = task_sheet.cell(row=row_idx + 2, column=task_name_col_idx)
                        
                        # Get font styling from row data
                        font_family = row.get('Font Family', 'Arial')
                        font_size = row.get('Font Size', 10)
                        font_color_hex = row.get('Font Color', '#000000').lstrip('#')
                        bg_color_hex = row.get('Background Color', '#FFFFFF').lstrip('#')
                        is_bold = row.get('Font Bold', 'No') == 'Yes'
                        is_italic = row.get('Font Italic', 'No') == 'Yes'
                        is_underline = row.get('Font Underline', 'No') == 'Yes'
                        
                        # Create and apply font
                        cell_font = Font(
                            name=font_family,
                            size=font_size,
                            bold=is_bold,
                            italic=is_italic,
                            underline='single' if is_underline else None,
                            color=font_color_hex
                        )
                        task_name_cell.font = cell_font
                        
                        # Apply background color (override status color for this cell)
                        if bg_color_hex != 'FFFFFF':  # Only apply if not default white
                            bg_fill = PatternFill(start_color=bg_color_hex,
                                                 end_color=bg_color_hex,
                                                 fill_type='solid')
                            task_name_cell.fill = bg_fill
                
                # Write resources
                if resources_data:
                    df_resources = pd.DataFrame(resources_data)
                    df_resources.to_excel(writer, sheet_name='Resources', index=False)
                    
                    resource_sheet = writer.sheets['Resources']
                    for col in resource_sheet.columns:
                        max_length = max(len(str(cell.value)) for cell in col)
                        resource_sheet.column_dimensions[col[0].column_letter].width = max_length + 2
                
                # Write status breakdown
                df_status = pd.DataFrame(status_data)
                df_status.to_excel(writer, sheet_name='Status Breakdown', index=False)
                df_status = writer.sheets['Status Breakdown']
                for col in df_status.columns:
                        max_length = max(len(str(cell.value)) for cell in col)
                        df_status.column_dimensions[col[0].column_letter].width = max_length + 2
                
                # Write dependency breakdown
                df_deps = pd.DataFrame(dep_data)
                df_deps.to_excel(writer, sheet_name='Dependencies', index=False)
                df_deps = writer.sheets['Dependencies']
                for col in df_deps.columns:
                        max_length = max(len(str(cell.value)) for cell in col)
                        df_deps.column_dimensions[col[0].column_letter].width = max_length + 2
                
                # Write baselines
                baselines = data_manager.get_all_baselines()
                if baselines:
                    for baseline in baselines:
                        baseline_data = []
                        for snapshot in baseline.snapshots:
                            baseline_data.append({
                                'Task ID': snapshot.task_id,
                                'Task Name': snapshot.task_name,
                                'WBS': snapshot.wbs,
                                'Start Date': snapshot.start_date.strftime(date_format_str),
                                'End Date': snapshot.end_date.strftime(date_format_str),
                                'Duration': snapshot.duration,
                                '% Complete': snapshot.percent_complete
                            })
                        
                        if baseline_data:
                            df_baseline = pd.DataFrame(baseline_data)
                            # Sanitize sheet name (Excel has 31 char limit and no special chars)
                            sheet_name = f"BL_{baseline.name[:25]}"  # Prefix with BL_ and limit length
                            df_baseline.to_excel(writer, sheet_name=sheet_name, index=False)
                            
                            baseline_sheet = writer.sheets[sheet_name]
                            for col in baseline_sheet.columns:
                                max_length = max(len(str(cell.value)) for cell in col)
                                baseline_sheet.column_dimensions[col[0].column_letter].width = max_length + 2
                    
                    # Write baseline metadata sheet
                    baseline_meta_data = []
                    for baseline in baselines:
                        baseline_meta_data.append({
                            'Baseline Name': baseline.name,
                            'Created Date': baseline.created_date.strftime(date_format_str),
                            'Tasks Captured': len(baseline.snapshots)
                        })
                    
                    df_baseline_meta = pd.DataFrame(baseline_meta_data)
                    df_baseline_meta.to_excel(writer, sheet_name='Baselines', index=False)
                    
                    baseline_meta_sheet = writer.sheets['Baselines']
                    for col in baseline_meta_sheet.columns:
                        max_length = max(len(str(cell.value)) for cell in col)
                        baseline_meta_sheet.column_dimensions[col[0].column_letter].width = max_length + 2
            
            return True
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    @staticmethod
    def import_from_excel(filepath: str) -> Tuple[DataManager, bool]:
        """Import project from Excel file with backward compatibility"""
        data_manager = DataManager()
        calendar_manager = CalendarManager()
        data_manager.calendar_manager = calendar_manager
        date_format_enum = DateFormat.YYYY_MM_DD # Default
        
        try:
            # Try to read summary for project name, project start date and date format
            try:
                df_summary = pd.read_excel(filepath, sheet_name='Summary')
                for _, row in df_summary.iterrows():
                    if row['Metric'] == 'Project Name':
                        data_manager.project_name = str(row['Value'])
                    elif row['Metric'] == 'Project Start Date':
                        start_date_str = str(row['Value'])
                        if start_date_str != 'N/A':
                            # Attempt to parse with default YYYY-MM-DD first
                            try:
                                data_manager.settings.project_start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                            except ValueError:
                                # Fallback to isoformat if previous fails
                                data_manager.settings.project_start_date = pd.to_datetime(start_date_str, dayfirst=False).to_pydatetime()
                    elif row['Metric'] == 'Date Format':
                        try:
                            date_format_enum = DateFormat(str(row['Value']))
                        except ValueError:
                            date_format_enum = DateFormat.YYYY_MM_DD # Fallback
                    elif row['Metric'] == 'Working Hours Start':
                        try:
                            wh_start = str(row['Value'])
                            if ':' in wh_start:
                                calendar_manager.working_hours_start = wh_start
                        except:
                            pass
                    elif row['Metric'] == 'Working Hours End':
                        try:
                            wh_end = str(row['Value'])
                            if ':' in wh_end:
                                calendar_manager.working_hours_end = wh_end
                        except:
                            pass
            
                # Recalculate hours per day based on imported times
                calendar_manager.hours_per_day = calendar_manager._calculate_hours_from_times()
                
                # Set calendar manager reference immediately so it's available
                data_manager.calendar_manager = calendar_manager
            except:
                pass
            
            date_format_str = Exporter._get_date_format_string(date_format_enum)
            
            # Read tasks
            df_tasks = pd.read_excel(filepath, sheet_name='Tasks')
            
            # Check if this is new format (with Level and Parent ID columns)
            has_hierarchy = 'Level' in df_tasks.columns and 'Parent ID' in df_tasks.columns
            has_is_summary = 'Is Summary' in df_tasks.columns
            has_is_milestone = 'Is Milestone' in df_tasks.columns # Check for is_milestone column
            
            for _, row in df_tasks.iterrows():
                # Parse predecessors with dependency types
                predecessors = []
                if pd.notna(row.get('Predecessors', '')) and str(row['Predecessors']).strip():
                    pred_str = str(row['Predecessors'])
                    
                    # Check if new format with dependency types
                    if '(' in pred_str:
                        # New format: "1 (Finish-to-Start), 2 (Start-to-Start)"
                        for pred_item in pred_str.split(','):
                            pred_item = pred_item.strip()
                            if '(' in pred_item:
                                # Extract ID and type
                                parts = pred_item.split('(')
                                pred_id = int(parts[0].strip())
                                dep_type_text_raw = parts[1].replace(')', '').strip()
                                dep_type_text = dep_type_text_raw
                                lag_days = 0

                                # Check for lag days
                                lag_match = re.search(r'([+-]?\d+)d?$', dep_type_text_raw)
                                if lag_match:
                                    lag_days = int(lag_match.group(1))
                                    dep_type_text = dep_type_text_raw[:lag_match.start()].strip() # Remove lag from dep_type_text

                                # Convert text to enum name
                                dep_type = None
                                for dt in DependencyType:
                                    if dt.value == dep_type_text:
                                        dep_type = dt.name
                                        break
                                
                                if dep_type is None:
                                    dep_type = DependencyType.FS.name
                                
                                predecessors.append((pred_id, dep_type, lag_days))
                            else:
                                # Fallback: just ID, assume FS
                                try:
                                    pred_id = int(pred_item.strip())
                                    predecessors.append((pred_id, DependencyType.FS.name, 0))
                                except:
                                    pass
                    else:
                        # Old format: just IDs "1,2,3"
                        for pred_id_str in pred_str.split(','):
                            try:
                                pred_id = int(pred_id_str.strip())
                                predecessors.append((pred_id, DependencyType.FS.name, 0))
                            except:
                                pass
                
                # Parse resources with allocation
                assigned_resources = []
                if pd.notna(row.get('Assigned Resources', '')) and str(row['Assigned Resources']).strip():
                    resources_str = str(row['Assigned Resources'])
                    for r_item in resources_str.split(','):
                        r_item = r_item.strip()
                        match = re.match(r'(.+?)\s*\((\d+)\s*%\)', r_item)
                        if match:
                            assigned_resources.append((match.group(1).strip(), int(match.group(2))))
                        else:
                            # Backward compatibility: assume 100% if not specified
                            assigned_resources.append((r_item, 100))
                
                # Get parent ID
                parent_id = None
                if has_hierarchy and pd.notna(row.get('Parent ID', '')):
                    try:
                        parent_id = int(row['Parent ID'])
                    except:
                        pass
                
                # Get is_summary flag
                is_summary = False
                if has_is_summary and pd.notna(row.get('Is Summary', '')):
                    is_summary = str(row['Is Summary']).lower() in ['yes', 'true', '1']
                
                # Clean task name (remove summary marker, but preserve indentation if present)
                task_name = str(row['Task Name']).strip()
                # Only remove the summary marker if it's at the very beginning
                if task_name.startswith('▶'):
                    task_name = task_name.lstrip('▶').strip()
                
                # Get is_milestone flag
                is_milestone = False
                if has_is_milestone and pd.notna(row.get('Is Milestone', '')):
                    is_milestone = str(row['Is Milestone']).lower() in ['yes', 'true', '1']
                
                start_date_val = row['Start Date']
                end_date_val = row['End Date']

                # Handle potential 'Milestone' string in End Date column for old exports
                if isinstance(end_date_val, str) and end_date_val.lower() == 'milestone':
                    end_date_val = start_date_val

                task = Task(
                    name=task_name,
                    start_date=datetime.strptime(str(start_date_val), date_format_str),
                    end_date=datetime.strptime(str(end_date_val), date_format_str),
                    percent_complete=int(row.get('% Complete', 0)),
                    predecessors=predecessors,
                    assigned_resources=assigned_resources,
                    notes=str(row.get('Notes', '')) if pd.notna(row.get('Notes', '')) else '',
                    task_id=int(row['Task ID']),
                    parent_id=parent_id,
                    is_summary=is_summary,
                    is_milestone=is_milestone,
                    wbs=row.get('WBS'),
                    schedule_type=ScheduleType(str(row.get('Schedule Mode', 'Auto Scheduled'))) if str(row.get('Schedule Mode', 'Auto Scheduled')) in [item.value for item in ScheduleType] else ScheduleType.AUTO_SCHEDULED,
                    # Font styling with backward compatibility
                    font_family=str(row.get('Font Family', 'Arial')) if pd.notna(row.get('Font Family')) else 'Arial',
                    font_size=int(row.get('Font Size', 10)) if pd.notna(row.get('Font Size')) else 10,
                    font_color=str(row.get('Font Color', '#000000')) if pd.notna(row.get('Font Color')) else '#000000',
                    background_color=str(row.get('Background Color', '#FFFFFF')) if pd.notna(row.get('Background Color')) else '#FFFFFF',
                    font_bold=(str(row.get('Font Bold', 'No')).lower() in ['yes', 'true', '1']) if pd.notna(row.get('Font Bold')) else False,
                    font_italic=(str(row.get('Font Italic', 'No')).lower() in ['yes', 'true', '1']) if pd.notna(row.get('Font Italic')) else False,
                    font_underline=(str(row.get('Font Underline', 'No')).lower() in ['yes', 'true', '1']) if pd.notna(row.get('Font Underline')) else False
                )
                data_manager.tasks.append(task)
            
            # After all tasks are loaded, generate WBS and then determine summary/milestone status
            data_manager._generate_wbs()

            # Auto-identify milestones and summary tasks
            for task in data_manager.tasks:
                # Identify milestones: duration of 0
                if task.get_duration(data_manager.settings.duration_unit, data_manager.calendar_manager) == 0:
                    task.is_milestone = True
                    task.end_date = task.start_date # Ensure end_date equals start_date for milestones

                # Identify summary tasks: if it has children
                if data_manager.get_child_tasks(task.id):
                    task.is_summary = True

            # If project_start_date was not loaded from summary, infer from earliest task
            if data_manager.settings.project_start_date is None and data_manager.tasks:
                earliest_start = min(task.start_date for task in data_manager.tasks)
                data_manager.settings.project_start_date = earliest_start

            # Set the default_date_format in data_manager.settings
            data_manager.settings.default_date_format = date_format_enum

            # Read resources if sheet exists
            try:
                df_resources = pd.read_excel(filepath, sheet_name='Resources')
                
                for _, row in df_resources.iterrows():
                    exceptions = []
                    if pd.notna(row.get('Exceptions', '')) and str(row['Exceptions']).strip():
                        exceptions = [x.strip() for x in str(row['Exceptions']).split(',')]
                    
                    resource = Resource(
                        name=str(row['Resource Name']),
                        max_hours_per_day=float(row.get('Max Hours/Day', 8.0)),
                        exceptions=exceptions
                    )
                    data_manager.resources.append(resource)
            except Exception as e:
                print(f"Error importing resources from Excel: {e}")
                import traceback
                traceback.print_exc()
            
            # Import baselines (with backward compatibility)
            try:
                from data_manager.models import Baseline, TaskSnapshot
                
                # Try to read baseline metadata sheet
                df_baseline_meta = pd.read_excel(filepath, sheet_name='Baselines')
                
                for _, meta_row in df_baseline_meta.iterrows():
                    baseline_name = str(meta_row['Baseline Name'])
                    created_date_str = str(meta_row['Created Date'])
                    
                    # Parse created date
                    try:
                        created_date = datetime.strptime(created_date_str, date_format_str)
                    except:
                        created_date = datetime.now()
                    
                    # Create baseline
                    baseline = Baseline(name=baseline_name, created_date=created_date)
                    
                    # Read baseline task data from corresponding sheet
                    sheet_name = f"BL_{baseline_name[:25]}"
                    try:
                        df_baseline_tasks = pd.read_excel(filepath, sheet_name=sheet_name)
                        
                        for _, task_row in df_baseline_tasks.iterrows():
                            snapshot = TaskSnapshot(
                                task_id=int(task_row['Task ID']),
                                task_name=str(task_row['Task Name']),
                                start_date=datetime.strptime(str(task_row['Start Date']), date_format_str),
                                end_date=datetime.strptime(str(task_row['End Date']), date_format_str),
                                duration=float(task_row['Duration']),
                                percent_complete=int(task_row['% Complete']),
                                wbs=str(task_row.get('WBS', ''))
                            )
                            baseline.add_task_snapshot(snapshot)
                        
                        data_manager.baselines.append(baseline)
                    except Exception as sheet_error:
                        print(f"Error reading baseline sheet {sheet_name}: {sheet_error}")
                        continue
                        
            except Exception as e:
                # Baselines sheet doesn't exist - backward compatibility
                print(f"No baselines found in Excel file (backward compatibility): {e}")
            
            return data_manager, True
            
        except Exception as e:
            print(f"Error importing from Excel: {e}")
            import traceback
            traceback.print_exc()
            return data_manager, False
    
    @staticmethod
    def export_to_json(data_manager: DataManager, calendar_manager: CalendarManager, 
                     filepath: str) -> bool:
        """Save project to JSON file with all enhanced features"""
        try:
            data = {
                'version': '2.2',
                'project_name': data_manager.project_name,
                'project_data': data_manager.save_to_dict(),
                'calendar_settings': calendar_manager.to_dict(),
                'project_settings': data_manager.settings.to_dict(), # Include project settings
                'saved_at': datetime.now().isoformat(),
                'metadata': {
                    'total_tasks': len(data_manager.tasks),
                    'summary_tasks': sum(1 for t in data_manager.tasks if t.is_summary),
                    'work_tasks': sum(1 for t in data_manager.tasks if not t.is_summary),
                    'total_resources': len(data_manager.resources),
                    'overall_completion': data_manager.get_overall_completion(),
                    'duration_unit': data_manager.settings.duration_unit.value
                }
            }
            
            with open(filepath, 'w') as f:
                json.dump(data, f, indent=2)
            
            return True
            
        except Exception as e:
            print(f"Error saving to JSON: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    @staticmethod
    def import_from_json(filepath: str) -> Tuple[DataManager, CalendarManager, bool]:
        """Load project from JSON file with backward compatibility"""
        data_manager = DataManager()
        calendar_manager = CalendarManager()
        
        try:
            with open(filepath, 'r') as f:
                data = json.load(f)
            
            # Check version for backward compatibility
            version = data.get('version', '1.0')
            
            # Load project name (new in v2.0)
            if 'project_name' in data:
                data_manager.project_name = data['project_name']
            elif 'project_data' in data and 'project_name' in data['project_data']:
                data_manager.project_name = data['project_data']['project_name']
            else:
                # Try to extract from filename
                import os
                filename = os.path.basename(filepath)
                data_manager.project_name = os.path.splitext(filename)[0].replace('_', ' ')
            
            # Load project settings (new in v2.2)
            if 'project_settings' in data:
                data_manager.settings.from_dict(data['project_settings'])

            # Load calendar settings (new in v2.0)
            if 'calendar_settings' in data:
                calendar_manager.from_dict(data['calendar_settings'])

            # Load project data
            data_manager.load_from_dict(data.get('project_data', {}))
            
            # Backward compatibility for project_start_date if not loaded from settings
            if data_manager.settings.project_start_date is None and data_manager.tasks:
                # If no project_start_date in settings, use the earliest task start date
                earliest_start = min(task.start_date for task in data_manager.tasks)
                data_manager.settings.project_start_date = earliest_start

            # Set calendar manager reference
            data_manager.calendar_manager = calendar_manager
            
            return data_manager, calendar_manager, True
            
        except Exception as e:
            print(f"Error loading from JSON: {e}")
            import traceback
            traceback.print_exc()
            return data_manager, calendar_manager, False
    
    @staticmethod
    def export_to_csv(data_manager: DataManager, filepath: str) -> bool:
        """Export tasks to CSV format (flat structure)"""
        try:
            tasks_data = []
            
            for task in sorted(data_manager.get_all_tasks(), key=lambda t: t.id):
                # Format predecessors
                pred_text = ', '.join([
                    f"{pred_id}{DependencyType[dep_type].value}{f'{lag_days:+d}d' if lag_days != 0 else ''}" 
                    for pred_id, dep_type, lag_days in task.predecessors
                ])
                
                tasks_data.append({
                    'ID': task.id,
                    'Name': task.name,
                    'Parent_ID': task.parent_id if task.parent_id else '',
                    'Is_Summary': task.is_summary,
                    'Start': task.start_date.strftime('%Y-%m-%d'),
                    'End': task.end_date.strftime('%Y-%m-%d'),
                    'Duration': task.duration,
                    'Complete': task.percent_complete,
                    'Status': task.get_status_text(),
                    'Predecessors': pred_text,
                    'Resources': ','.join(task.assigned_resources),
                    'Notes': task.notes
                })
            
            df = pd.DataFrame(tasks_data)
            df.to_csv(filepath, index=False)
            
            return True
            
        except Exception as e:
            print(f"Error exporting to CSV: {e}")
            return False
    
    @staticmethod
    def export_project_report(data_manager: DataManager, filepath: str) -> bool:
        """Export comprehensive project report to Excel"""
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Project Overview
                overview_data = {
                    'Item': [
                        'Project Name',
                        'Start Date',
                        'End Date',
                        'Duration (days)',
                        'Overall Completion (%)',
                        'Total Tasks',
                        'Summary Tasks',
                        'Work Tasks',
                        'Resources',
                        'Report Date'
                    ],
                    'Value': [
                        data_manager.project_name,
                        data_manager.get_project_start_date().strftime('%Y-%m-%d') 
                            if data_manager.get_project_start_date() else 'N/A',
                        data_manager.get_project_end_date().strftime('%Y-%m-%d')
                            if data_manager.get_project_end_date() else 'N/A',
                        (data_manager.get_project_end_date() - data_manager.get_project_start_date()).days + 1
                            if data_manager.get_project_start_date() and data_manager.get_project_end_date() else 'N/A',
                        f"{data_manager.get_overall_completion():.1f}",
                        len(data_manager.tasks),
                        sum(1 for t in data_manager.tasks if t.is_summary),
                        sum(1 for t in data_manager.tasks if not t.is_summary),
                        len(data_manager.resources),
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    ]
                }
                pd.DataFrame(overview_data).to_excel(writer, sheet_name='Overview', index=False)
                
                # Task Status Summary
                status_summary = {}
                for task in data_manager.tasks:
                    if not task.is_summary:
                        status = task.get_status_text()
                        status_summary[status] = status_summary.get(status, 0) + 1
                
                status_data = {
                    'Status': list(status_summary.keys()),
                    'Count': list(status_summary.values()),
                    'Percentage': [f"{(count/sum(status_summary.values())*100):.1f}" 
                                  for count in status_summary.values()]
                }
                pd.DataFrame(status_data).to_excel(writer, sheet_name='Status Summary', index=False)
                
                # Critical Tasks (Overdue)
                critical_tasks = []
                for task in data_manager.tasks:
                    if not task.is_summary and task.get_status_text() == 'Overdue':
                        critical_tasks.append({
                            'ID': task.id,
                            'Task': task.name,
                            'End Date': task.end_date.strftime('%Y-%m-%d'),
                            'Days Overdue': (datetime.now() - task.end_date).days,
                            'Complete': f"{task.percent_complete}%",
                            'Resources': ', '.join(task.assigned_resources)
                        })
                
                if critical_tasks:
                    pd.DataFrame(critical_tasks).to_excel(writer, sheet_name='Critical Tasks', index=False)
                
                # Resource Utilization
                allocation = data_manager.get_resource_allocation()
                resource_data = []
                for resource in data_manager.resources:
                    alloc = allocation.get(resource.name, {})
                    resource_data.append({
                        'Resource': resource.name,
                        'Max Hours/Day': resource.max_hours_per_day,
                        'Total Hours': f"{alloc.get('total_hours', 0):.1f}",
                        'Tasks': alloc.get('tasks_assigned', 0),
                        'Utilization': f"{(alloc.get('total_hours', 0) / (resource.max_hours_per_day * 260)):.1%}"
                    })
                
                pd.DataFrame(resource_data).to_excel(writer, sheet_name='Resources', index=False)
                
                # Dependency Analysis
                dep_analysis = {
                    'Finish-to-Start': 0,
                    'Start-to-Start': 0,
                    'Finish-to-Finish': 0,
                    'Start-to-Finish': 0
                }
                
                for task in data_manager.tasks:
                    for _, dep_type in task.predecessors:
                        try:
                            dep_name = DependencyType[dep_type].value
                            dep_analysis[dep_name] = dep_analysis.get(dep_name, 0) + 1
                        except:
                            pass
                
                dep_data = {
                    'Dependency Type': list(dep_analysis.keys()),
                    'Count': list(dep_analysis.values())
                }
                pd.DataFrame(dep_data).to_excel(writer, sheet_name='Dependencies', index=False)
            
            return True
            
        except Exception as e:
            print(f"Error exporting project report: {e}")
            import traceback
            traceback.print_exc()
            return False